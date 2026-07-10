import sys
import random
import requests
import openpyxl
import io

BASE = "http://localhost:8000"

def req(method, path, token=None, **kw):
    headers = kw.pop("headers", {})
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return requests.request(method, f"{BASE}{path}", headers=headers, **kw)

def check(cond, msg):
    if not cond:
        print(f"FAIL: {msg}")
        sys.exit(1)
    print(f"OK: {msg}")

def generate_excel_bytes(suffix):
    wb = openpyxl.Workbook()
    
    # 1. Sheet: Classes & Sections
    ws_classes = wb.active
    ws_classes.title = "Classes & Sections"
    ws_classes.append(["Class Name", "Section Name"])
    ws_classes.append([f"Grade 9-{suffix}", "A"])
    ws_classes.append([f"Grade 9-{suffix}", "B"])
    ws_classes.append([f"Grade 10-{suffix}", "A"])
    
    # 2. Sheet: Resources
    ws_resources = wb.create_sheet(title="Resources")
    ws_resources.append(["Resource Name", "Type", "Capacity"])
    ws_resources.append([f"Chemistry Lab-{suffix}", "Lab", 30])
    ws_resources.append([f"Main Hall-{suffix}", "Auditorium", 100])
    
    # 3. Sheet: Subjects
    ws_subjects = wb.create_sheet(title="Subjects")
    ws_subjects.append(["Subject Name", "Weekly Hours", "Resource Name"])
    ws_subjects.append([f"Mathematics-{suffix}", 5, ""])
    ws_subjects.append([f"Chemistry-{suffix}", 4, f"Chemistry Lab-{suffix}"])
    ws_subjects.append([f"Physical Education-{suffix}", 2, f"Main Hall-{suffix}"])
    
    # 4. Sheet: Teachers
    ws_teachers = wb.create_sheet(title="Teachers")
    ws_teachers.append(["Teacher Name", "Email", "Password", "Department", "Max Weekly Hours", "Subjects"])
    ws_teachers.append([f"Teacher One-{suffix}", f"t1.{suffix}@school.com", "Password@123", "Science", 30, f"Mathematics-{suffix}, Chemistry-{suffix}"])
    ws_teachers.append([f"Teacher Two-{suffix}", f"t2.{suffix}@school.com", "Password@123", "Arts", 20, f"Physical Education-{suffix}"])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()

def main():
    suffix = random.randint(10000, 99999)
    from app.database import SessionLocal, Base, engine
    from app import models
    from app.auth import hash_password
    Base.metadata.create_all(bind=engine)
    
    db = SessionLocal()
    # Create super_admin
    sa_email = f"bulk.sa.{suffix}@example.com"
    sa = models.User(
        name="Bulk SuperAdmin",
        email=sa_email,
        hashed_password=hash_password("pass1234"),
        role=models.RoleEnum.super_admin
    )
    db.add(sa)
    db.commit()
    db.close()

    # Login as Super Admin
    r = req("POST", "/auth/login", json={"email": sa_email, "password": "pass1234"})
    check(r.status_code == 200, "super_admin login")
    sa_token = r.json()["access_token"]

    # Create a School
    r = req("POST", "/schools", token=sa_token, json={
        "name": f"Bulk School {suffix}",
        "periods_per_day": 8,
        "working_days": 5
    })
    check(r.status_code == 201, "create school")
    school_id = r.json()["id"]

    # Generate Excel payload
    excel_data = generate_excel_bytes(suffix)

    # Upload Excel file
    files = {"file": ("test_bulk.xlsx", excel_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = req("POST", f"/bulk/upload?school_id={school_id}", token=sa_token, files=files)
    check(r.status_code == 200, "upload bulk Excel file")
    
    res_data = r.json()
    check(res_data["success"] is True, "success status in upload response")
    
    stats = res_data["stats"]
    check(stats["classes_created"] == 2, "classes created count matches")
    check(stats["sections_created"] == 3, "sections created count matches")
    check(stats["resources_created"] == 2, "resources created count matches")
    check(stats["subjects_created"] == 3, "subjects created count matches")
    check(stats["teachers_created"] == 2, "teachers created count matches")

    # Fetch classes to verify
    r = req("GET", f"/classes?limit=100", token=sa_token)
    classes = [c for c in r.json()["items"] if c["school_id"] == school_id]
    check(len(classes) == 2, "verify 2 classes are retrieved via API")

    # Fetch subjects to verify
    r = req("GET", f"/subjects?limit=100", token=sa_token)
    subjects = [s for s in r.json()["items"] if s["school_id"] == school_id]
    check(len(subjects) == 3, "verify 3 subjects are retrieved via API")
    # Verify resource mapping
    chem_subject = next(s for s in subjects if f"Chemistry-{suffix}" in s["name"])
    check(chem_subject["resource_id"] is not None, "verify Chemistry subject is linked to a resource")

    # Fetch teachers to verify
    r = req("GET", f"/teachers?limit=100", token=sa_token)
    teachers = [t for t in r.json()["items"] if t["school_id"] == school_id]
    check(len(teachers) == 2, "verify 2 teachers are retrieved via API")
    t1 = next(t for t in teachers if f"t1.{suffix}" in t["email"])
    check(len(t1["subject_ids"]) == 2, "verify teacher qualified subjects are mapped")

    # Generate timetable using the uploaded data
    r = req("POST", "/timetables/generate", token=sa_token, json={"school_id": school_id, "time_limit_seconds": 10})
    check(r.status_code == 200, "run OR-Tools scheduling solver on uploaded data")
    check(r.json()["slots_created"] > 0, "verify schedule slots were generated successfully")

    print("\nALL BULK UPLOAD INTEGRATION TESTS PASSED SUCCESSFULLY!")

if __name__ == "__main__":
    main()
