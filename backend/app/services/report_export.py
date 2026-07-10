"""
Export helpers for the Reports module (Reports, Phase 7).

Two shapes are supported:
  - a plain table (headers + rows) -> used by Teacher Workload, Subject
    Coverage, Resource Usage, and Leave Summary.
  - a day/period grid -> used by the Timetables report (mirrors the
    section/teacher grid views already in the Timetable Viewer UI).

Kept deliberately generic/dependency-light: reportlab for PDF, openpyxl for
Excel. Both return raw bytes so the router can stream them back without
touching the filesystem.
"""
import io
from datetime import date, datetime, time

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def _stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    return str(value)


def rows_to_pdf(*, title: str, subtitle: str, headers: list[str], rows: list[list]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 12))

    data = [headers] + [[_stringify(c) for c in r] for r in rows]
    if len(rows) == 0:
        data.append(["No data for the selected filters." if headers else ""] + [""] * (len(headers) - 1))

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def rows_to_xlsx(*, title: str, headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] if title else "Report"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_stringify(value) if isinstance(value, (date, datetime, time)) else value)

    for col_idx, h in enumerate(headers, start=1):
        max_len = max([len(str(h))] + [len(_stringify(r[col_idx - 1])) for r in rows] or [0])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def grid_to_pdf(*, title: str, subtitle: str, periods: list[int], working_days: int, cells: dict) -> bytes:
    """cells: {(day_of_week, period): "line1\\nline2"} -> multi-line cell text."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter), topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 12))

    header = ["Period"] + DAY_NAMES[:working_days]
    data = [header]
    for p in periods:
        row = [f"Period {p}"]
        for d in range(working_days):
            row.append(cells.get((d, p), "—"))
        data.append(row)

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#f8fafc")),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def grid_to_xlsx(*, title: str, periods: list[int], working_days: int, cells: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] if title else "Timetable"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.cell(row=1, column=1, value="Period").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    for i, day_name in enumerate(DAY_NAMES[:working_days], start=2):
        cell = ws.cell(row=1, column=i, value=day_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, p in enumerate(periods, start=2):
        ws.cell(row=row_idx, column=1, value=f"Period {p}").font = Font(bold=True)
        for d in range(working_days):
            text = cells.get((d, p), "")
            cell = ws.cell(row=row_idx, column=d + 2, value=text.replace("\n", " / ") if text else "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 12
    for i in range(2, 2 + working_days):
        ws.column_dimensions[get_column_letter(i)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
