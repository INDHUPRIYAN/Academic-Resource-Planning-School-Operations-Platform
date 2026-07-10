import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.auth import require_roles, hash_password
from app import models
from app.crud_factory import log_action

router = APIRouter(prefix="/bulk", tags=["bulk"])
ADMIN_ROLES = (models.RoleEnum.super_admin, models.RoleEnum.school_admin)

@router.get("/template")
def download_template(user: models.User = Depends(require_roles(*ADMIN_ROLES))):
    """
    Generates a beautifully styled .xlsx template file containing 4 sheets:
    - Classes & Sections
    - Resources
    - Subjects
    - Teachers
    """
    wb = openpyxl.Workbook()
    
    # Fonts & Colors
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    normal_font = Font(name="Segoe UI", size=10)
    note_font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
    example_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Light slate for example rows
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # 1. Sheet: Classes & Sections
    ws_classes = wb.active
    ws_classes.title = "Classes & Sections"
    ws_classes.append(["Class Name", "Section Name"])
    ws_classes.append(["Grade 9", "A"])
    ws_classes.append(["Grade 9", "B"])
    ws_classes.append(["Grade 10", "A"])
    
    # 2. Sheet: Resources
    ws_resources = wb.create_sheet(title="Resources")
    ws_resources.append(["Resource Name", "Type", "Capacity"])
    ws_resources.append(["Chemistry Lab", "Lab", 30])
    ws_resources.append(["Main Hall", "Auditorium", 100])
    
    # 3. Sheet: Subjects
    ws_subjects = wb.create_sheet(title="Subjects")
    ws_subjects.append(["Subject Name", "Weekly Hours", "Resource Name"])
    ws_subjects.append(["Mathematics", 5, ""])
    ws_subjects.append(["Chemistry", 4, "Chemistry Lab"])
    ws_subjects.append(["Physical Education", 2, "Main Hall"])
    
    # 4. Sheet: Teachers
    ws_teachers = wb.create_sheet(title="Teachers")
    ws_teachers.append(["Teacher Name", "Email", "Password", "Department", "Max Weekly Hours", "Subjects"])
    ws_teachers.append(["John Doe", "john.doe@school.com", "Password@123", "Science", 30, "Mathematics, Chemistry"])
    ws_teachers.append(["Jane Smith", "jane.smith@school.com", "Password@123", "Arts", 20, "Physical Education"])

    # Style worksheets
    for ws in wb.worksheets:
        ws.views.sheetView[0].showGridLines = True
        ws.row_dimensions[1].height = 28
        
        # Headers styling
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            
        # Set widths nicely
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 15)
            
        # Example data styling
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = normal_font
                cell.fill = example_fill
                cell.alignment = left_align
                cell.border = thin_border

    # Save to a memory stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=eduflow_bulk_template.xlsx"}
    )

@router.post("/upload")
async def upload_bulk_data(
    school_id: int | None = Query(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: models.User = Depends(require_roles(*ADMIN_ROLES))
):
    """
    Parses and imports Classes, Sections, Resources, Subjects, and Teachers from an uploaded Excel file.
    Performs full transaction validation. Rolled back on any error.
    """
    # Scope check
    target_school_id = user.school_id if user.role == models.RoleEnum.school_admin else school_id
    if not target_school_id:
        raise HTTPException(status_code=400, detail="school_id parameter is required for super_admin")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file format: {str(e)}")

    # We perform database operations in an explicit savepoint/transaction block so we can rollback if anything fails
    # Let's initialize stats
    stats = {
        "classes_created": 0,
        "sections_created": 0,
        "resources_created": 0,
        "subjects_created": 0,
        "teachers_created": 0,
    }

    try:
        # Load sheets
        sheet_names = wb.sheetnames
        
        # 1. Import Classes & Sections
        classes_map = {} # name -> class object
        sections_map = {} # (class_name, section_name) -> section object
        
        if "Classes & Sections" in sheet_names:
            ws = wb["Classes & Sections"]
            # Iterate rows skipping header
            for row in list(ws.rows)[1:]:
                # skip empty rows
                if not any(cell.value for cell in row):
                    continue
                class_name = str(row[0].value or "").strip()
                section_name = str(row[1].value or "").strip()
                
                if not class_name:
                    continue
                
                # Find or create class
                if class_name not in classes_map:
                    cls_obj = db.query(models.Class).filter(
                        models.Class.school_id == target_school_id,
                        models.Class.name == class_name
                    ).first()
                    if not cls_obj:
                        cls_obj = models.Class(school_id=target_school_id, name=class_name)
                        db.add(cls_obj)
                        db.flush()
                        stats["classes_created"] += 1
                    classes_map[class_name] = cls_obj
                
                cls_obj = classes_map[class_name]
                
                if section_name:
                    sec_key = (class_name, section_name)
                    if sec_key not in sections_map:
                        sec_obj = db.query(models.Section).filter(
                            models.Section.class_id == cls_obj.id,
                            models.Section.name == section_name
                        ).first()
                        if not sec_obj:
                            sec_obj = models.Section(class_id=cls_obj.id, name=section_name)
                            db.add(sec_obj)
                            db.flush()
                            stats["sections_created"] += 1
                        sections_map[sec_key] = sec_obj

        # 2. Import Resources
        resources_map = {} # name -> resource object
        if "Resources" in sheet_names:
            ws = wb["Resources"]
            for row in list(ws.rows)[1:]:
                if not any(cell.value for cell in row):
                    continue
                res_name = str(row[0].value or "").strip()
                res_type = str(row[1].value or "").strip() or None
                capacity_val = row[2].value
                try:
                    capacity = int(capacity_val) if capacity_val is not None else None
                except ValueError:
                    capacity = None
                
                if not res_name:
                    continue
                
                res_obj = db.query(models.Resource).filter(
                    models.Resource.school_id == target_school_id,
                    models.Resource.name == res_name
                ).first()
                
                if not res_obj:
                    res_obj = models.Resource(
                        school_id=target_school_id,
                        name=res_name,
                        type=res_type,
                        capacity=capacity
                    )
                    db.add(res_obj)
                    db.flush()
                    stats["resources_created"] += 1
                resources_map[res_name] = res_obj

        # 3. Import Subjects
        subjects_map = {} # name -> subject object
        if "Subjects" in sheet_names:
            ws = wb["Subjects"]
            for row in list(ws.rows)[1:]:
                if not any(cell.value for cell in row):
                    continue
                sub_name = str(row[0].value or "").strip()
                hours_val = row[1].value
                try:
                    weekly_hours = int(hours_val) if hours_val is not None else 1
                except ValueError:
                    weekly_hours = 1
                res_name = str(row[2].value or "").strip()
                
                if not sub_name:
                    continue
                
                res_id = None
                if res_name:
                    # Resolve resource
                    res_obj = resources_map.get(res_name)
                    if not res_obj:
                        res_obj = db.query(models.Resource).filter(
                            models.Resource.school_id == target_school_id,
                            models.Resource.name == res_name
                        ).first()
                    if res_obj:
                        res_id = res_obj.id
                    else:
                        raise ValueError(f"Resource '{res_name}' specified for subject '{sub_name}' does not exist.")

                sub_obj = db.query(models.Subject).filter(
                    models.Subject.school_id == target_school_id,
                    models.Subject.name == sub_name
                ).first()
                
                if not sub_obj:
                    sub_obj = models.Subject(
                        school_id=target_school_id,
                        name=sub_name,
                        weekly_hours=weekly_hours,
                        resource_id=res_id
                    )
                    db.add(sub_obj)
                    db.flush()
                    stats["subjects_created"] += 1
                else:
                    # Update hours and resource if specified
                    sub_obj.weekly_hours = weekly_hours
                    if res_id:
                        sub_obj.resource_id = res_id
                subjects_map[sub_name] = sub_obj

        # 4. Import Teachers
        if "Teachers" in sheet_names:
            ws = wb["Teachers"]
            for row in list(ws.rows)[1:]:
                if not any(cell.value for cell in row):
                    continue
                t_name = str(row[0].value or "").strip()
                t_email = str(row[1].value or "").strip()
                t_pwd = str(row[2].value or "").strip() or "Eduflow@123"
                t_dept = str(row[3].value or "").strip() or None
                hours_val = row[4].value
                try:
                    max_hours = int(hours_val) if hours_val is not None else 30
                except ValueError:
                    max_hours = 30
                t_subjects_raw = str(row[5].value or "").strip()
                
                if not t_name or not t_email:
                    continue
                
                # Check User / Teacher existence
                user_obj = db.query(models.User).filter(models.User.email == t_email).first()
                if user_obj:
                    # User exists, check if teacher profile exists
                    t_obj = db.query(models.Teacher).filter(models.Teacher.user_id == user_obj.id).first()
                    if not t_obj:
                        t_obj = models.Teacher(
                            user_id=user_obj.id,
                            school_id=target_school_id,
                            department=t_dept,
                            max_weekly_hours=max_hours
                        )
                        db.add(t_obj)
                        db.flush()
                        stats["teachers_created"] += 1
                else:
                    # Create new user & teacher
                    user_obj = models.User(
                        school_id=target_school_id,
                        name=t_name,
                        email=t_email,
                        hashed_password=hash_password(t_pwd),
                        role=models.RoleEnum.teacher
                    )
                    db.add(user_obj)
                    db.flush()
                    
                    t_obj = models.Teacher(
                        user_id=user_obj.id,
                        school_id=target_school_id,
                        department=t_dept,
                        max_weekly_hours=max_hours
                    )
                    db.add(t_obj)
                    db.flush()
                    stats["teachers_created"] += 1
                
                # Link subjects if provided
                if t_subjects_raw and t_obj:
                    subj_names = [n.strip() for n in t_subjects_raw.split(",") if n.strip()]
                    linked_subjects = []
                    for sname in subj_names:
                        sobj = subjects_map.get(sname)
                        if not sobj:
                            sobj = db.query(models.Subject).filter(
                                models.Subject.school_id == target_school_id,
                                models.Subject.name == sname
                            ).first()
                        if sobj:
                            linked_subjects.append(sobj)
                        else:
                            raise ValueError(f"Subject '{sname}' specified for teacher '{t_name}' does not exist.")
                    t_obj.subjects = linked_subjects

        db.commit()
        log_action(db, user.id, "bulk_upload", f"school_id={target_school_id} stats={stats}")
        return {
            "success": True,
            "message": "Bulk import completed successfully.",
            "stats": stats
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Import failed during database transaction. All changes rolled back. Error: {str(e)}"
        )
